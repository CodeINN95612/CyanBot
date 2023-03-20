from . import config, globals, storage
import os
import json
import datetime
import discord

import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment


async def setUp():

    if not os.path.exists(globals.DATA_DIR):
        os.makedirs(globals.DATA_DIR)

    if not os.path.exists(globals.SERIES_FILE):
        with open(globals.SERIES_FILE, 'w') as f:
            json.dump({}, f)

    if not os.path.exists(globals.MSG_FILE):
        with open(globals.MSG_FILE, 'w') as f:
            json.dump([], f)

    if not os.path.exists(globals.ALLOWED_FILE):
        with open(globals.ALLOWED_FILE, 'w') as f:
            json.dump([], f)


async def scanMessages(discordClient: discord.Client):

    scanDays = config.config["scanDays"]
    testChannel = discordClient.get_channel(int(config.config["testChannel"]))
    if testChannel:
        await testChannel.send(f"Verificando mensagens dos últimos {scanDays} dias")

    messages = []

    channelId = config.config["submitChannel"]
    dateNow = datetime.datetime.utcnow()
    dateThen = dateNow - datetime.timedelta(days=scanDays)
    updateChannel = discordClient.get_channel(int(channelId))
    server_id = updateChannel.guild.id

    async for message in updateChannel.history(after=dateThen):

        if not isinstance(message, discord.Message) or message.author.bot:
            continue

        authorId = message.author.id
        authorName = message.author.display_name
        date = message.created_at.timestamp()
        content = message.content.lower()
        data = {
            "serverId": str(server_id),
            "authorId": str(authorId),
            "authorName": authorName,
            "date": date,
            "content": content
        }
        messages.append(data)

    validMessages = _validateMessages(messages)
    _parseOverwriteMessages(validMessages, globals.MSG_FILE)

    if testChannel:
        await testChannel.send(f"{len(validMessages)} mensagens verificadas")


def getMessages():
    with open(globals.MSG_FILE, 'r') as f:
        return json.load(f)


def insertMessage(message) -> bool:
    valid, obj = _validateMessage(message)
    if not valid:
        return False
    _parseAppendMessage(obj, globals.MSG_FILE)
    return True


def getSeries():
    with open(globals.SERIES_FILE, 'r') as f:
        return json.load(f)


def upsertSerie(sId, chId):
    with open(globals.SERIES_FILE, 'r') as f:
        dataFile = json.load(f)

    dataFile[sId] = chId

    with open(globals.SERIES_FILE, 'w') as f:
        json.dump(dataFile, f, indent=4)


def addAllowed(name):
    with open(globals.ALLOWED_FILE, 'r') as f:
        dataFile = json.load(f)

    dataFile.append(name)

    with open(globals.ALLOWED_FILE, 'w') as f:
        json.dump(dataFile, f, indent=4)


def isAllowed(name):
    with open(globals.ALLOWED_FILE, 'r') as f:
        dataFile = json.load(f)

    return name in [s.lower() for s in dataFile]


def _validateMessages(messages):

    valid_messages = []

    for message in messages:
        valid, obj = _validateMessage(message)
        if not valid:
            continue

        valid_messages.append(obj)

    return valid_messages


def _validateMessage(message):
    content = message['content']
    words = content.split()
    words = [word for word in words if '@' not in word]

    if len(words) < 3:
        return (False, {})

    try:
        numbers = float(words[-1])
    except:
        return (False, {})

    role_name = words[-2]
    role_values = [r['name']
                   for r in config.config["roles"] if r['name'].lower() == role_name]

    if len(role_values) == 0:
        return (False, {})

    serie = " ".join(words[:-2])
    if not storage.isAllowed(serie):
        return (False, {})

    return (True, {
        "serverId": message["serverId"],
        "authorId": message["authorId"],
        "authorName": message["authorName"],
        "date": message["date"],
        "serie": " ".join(words[:-2]),
        "function": role_values[0],
        "chapter": numbers
    })


def _parseAppendMessage(message, filename):
    with open(filename, 'r') as f:
        messages = json.load(f)
    messages.append(message)
    with open(filename, 'w') as f:
        json.dump(messages, f, indent=4)


def _parseAppendMessages(messages, filename):
    with open(filename, 'r') as file:
        existing_data = json.load(file)

    existing_data.extend(messages)

    with open(filename, 'w') as file:
        json.dump(existing_data, file, indent=4)


def _parseOverwriteMessages(messages, filename):
    with open(filename, 'w') as file:
        json.dump(messages, file, indent=4)


def IntoSpreadsheet(discordMessage: discord.Message, filename: str):
    with open(globals.MSG_FILE, 'r') as f:
        messagesFile = json.load(f)

    # filtrar servidor
    serverId = str(discordMessage.guild.id)
    now = datetime.datetime.now()
    messages = []
    for msg in messagesFile:
        if msg["serverId"] != serverId:
            continue
        date = datetime.datetime.fromtimestamp(msg["date"])
        if date.month == now.month and date.year == now.year:
            messages.append(msg)

    # construir nueva data
    data = {}
    def numColName(n: str) -> str: return f"num_{n}"
    def valColName(n: str) -> str: return f"val_{n}"
    for msg in messages:
        id = msg["authorId"]
        if id not in data:
            data[id] = {
                "author": msg["authorName"],
                "totalWork": 0,
                "totalValue": 0.0
            }
            for role in config.config["roles"]:
                role = role["name"]
                data[id][valColName(role)] = 0.0
                data[id][numColName(role)] = 0

        role = [role for role in config.config["roles"]
                if role["name"].lower() == msg["function"].lower()][0]
        roleName = role["name"]
        roleValue = role["value"]
        data[id]["totalWork"] += 1
        data[id]["totalValue"] += roleValue
        data[id][valColName(roleName)] += roleValue
        data[id][numColName(roleName)] += 1

    # Spreadsheet

    wb = openpyxl.Workbook()
    ws = wb.active

    # Crear un objeto de formato para el título de la columna
    titulo_formato = openpyxl.styles.NamedStyle(name="titulo_formato")
    titulo_formato.font = Font(bold=True, size=14)
    titulo_formato.fill = PatternFill(
        start_color="FFCC99", end_color="FFCC99", fill_type="solid")
    titulo_formato.border = Border(left=Side(style="thin"), right=Side(
        style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    titulo_formato.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)

    ws["A1"] = "DATE:"
    ws["B1"] = now.strftime("%d-%m-%Y")

    filaInicial = 4

    headers = ["User ID", "User", "Total Quantity", "Total Value"]
    for role in config.config["roles"]:
        headers.append("Q." + role["name"])
        headers.append("V. " + role["name"])

    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=filaInicial, column=i, value=header)
        cell.style = titulo_formato
        cell.alignment = openpyxl.styles.Alignment(
            wrapText=True, horizontal='center', vertical='center')
        ws.column_dimensions[cell.column_letter].width = 20

    for i, id in enumerate(data, start=1):
        row = i + filaInicial
        cell = data[id]
        ws.cell(row=row, column=1, value=id)
        ws.cell(row=row, column=2, value=cell["author"])
        ws.cell(row=row, column=3, value=cell["totalWork"])
        ws.cell(row=row, column=4, value=cell["totalValue"])
        col = 5
        for role in config.config["roles"]:
            roleName = role["name"]
            ws.cell(row=row, column=col, value=cell[numColName(roleName)])
            col += 1
            ws.cell(row=row, column=col, value=cell[valColName(roleName)])
            col += 1

    # Aplicar bordes
    for row in ws.iter_rows(min_row=1, min_col=1, max_col=25):
        for cell in row:
            if cell.value:
                cell.border = titulo_formato.border

    # Guardar archivo de Excel
    archivo = "./data/" + config.config["updateSpreadsheetName"] + '.xlsx'
    wb.save(archivo)

    # Enviar archivo como mensaje embebido en Discord
    with open(archivo, 'rb') as f:
        file = discord.File(f, filename=filename)
    return file
