from . import config
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.styles import numbers


def numColName(n: str) -> str: return f"num_{n}"
def valColName(n: str) -> str: return f"val_{n}"


def createWorkbook(archivo, data, roles, series, date, ago):
    wb = openpyxl.Workbook()

    _addData(wb, data, date, ago)
    _addRoles(wb, roles)
    _addSeries(wb, series)

    # Guardar archivo de Excel
    if os.path.exists(archivo):
        os.remove(archivo)
    wb.save(archivo)


def _addData(wb, data, date, ago):
    ws = wb.active
    ws.title = "Data"

    ws["A1"] = "FROM:"
    _styleCell(ws["A1"], BLACK, 14, WHITE, True)

    ws["B1"] = ago.strftime("%d-%m-%Y")
    _styleCell(ws["B1"], LIGHT_BLUE, 14, BLACK, False)

    ws["C1"] = "TO:"
    _styleCell(ws["C1"], BLACK, 14, WHITE, True)

    ws["D1"] = date.strftime("%d-%m-%Y")
    _styleCell(ws["D1"], LIGHT_BLUE, 14, BLACK, False)

    filaInicial = 4

    headers = ["#", "User ID", "User",
               "Total Quantity", "Total Value", "Roles"]
    for role in config.config["roles"]:
        headers.append(role["name"])
        # headers.append("V. " + role["name"])

    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=filaInicial, column=i, value=header)
        if (header == "Total Quantity"):
            _styleHeaderGreen(cell)
        elif (header == "Total Value"):
            _styleHeaderRed(cell)
        else:
            _styleHeaderNormal(cell)
        cell.alignment = openpyxl.styles.Alignment(
            wrapText=True, horizontal='center', vertical='center')
        ws.column_dimensions[cell.column_letter].width = 10 if i == 1 else 20

    for i, id in enumerate(data, start=1):
        row = i + filaInicial
        cell = data[id]

        _styleNormal(ws.cell(row=row, column=1, value=i))
        _styleNormal(ws.cell(row=row, column=2, value=id))
        _styleNormal(ws.cell(row=row, column=3, value=cell["author"]))
        _styleNormal(ws.cell(row=row, column=4,
                             value=cell["totalWork"])).number_format = numbers.FORMAT_NUMBER
        _styleNormal(ws.cell(row=row, column=5,
                             value=cell["totalValue"])).number_format = numbers.FORMAT_CURRENCY_USD
        _styleNormal(ws.cell(row=row, column=6, value=cell["roles"]))
        col = 7
        for role in config.config["roles"]:
            roleName = role["name"]

            cellr = ws.cell(row=row, column=col,
                            value=cell[numColName(roleName)])
            _styleNormal(cellr).number_format = numbers.FORMAT_NUMBER
            col += 1

            # cellr = ws.cell(row=row, column=col,
            #                 value=cell[valColName(roleName)])
            # _styleNormal(cellr).number_format = numbers.FORMAT_CURRENCY_USD
            # col += 1


def _addRoles(wb, roles):
    sheet = wb.create_sheet("Roles")

    sheet['A1'] = "#"
    _styleHeaderNormal(sheet['A1'])

    sheet['B1'] = "Role Name"
    _styleHeaderNormal(sheet['B1'])

    sheet['C1'] = "Value"
    _styleHeaderNormal(sheet['C1'])

    # loop through the array and add each role to a new row, along with its row number
    for i, role in enumerate(roles):
        row_number = i + 2  # add 2 because we started at row 2

        sheet.cell(row=row_number, column=1).value = row_number - 1
        _styleNormal(sheet.cell(row=row_number, column=1))

        sheet.cell(row=row_number, column=2).value = role["name"]
        _styleNormal(sheet.cell(row=row_number, column=2))

        sheet.cell(row=row_number, column=3).value = role["value"]
        _styleNormal(sheet.cell(row=row_number, column=3))
        sheet.cell(
            row=row_number, column=3).number_format = numbers.FORMAT_CURRENCY_USD

    sheet.column_dimensions['A'].width = 2
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 10


def _addSeries(wb, series):
    sheet = wb.create_sheet("Series")

    sheet['A1'] = "#"
    _styleHeaderNormal(sheet['A1'])

    sheet['B1'] = "Series"
    _styleHeaderNormal(sheet['B1'])

    for i, name in enumerate(series):
        row_number = i + 2  # add 2 because we started at row 2
        sheet.cell(row=row_number, column=1).value = row_number - 1
        _styleNormal(sheet.cell(row=row_number, column=1))

        sheet.cell(row=row_number, column=2).value = name
        _styleNormal(sheet.cell(row=row_number, column=2))

    sheet.column_dimensions['A'].width = 2
    sheet.column_dimensions['B'].width = 30


GREEN = '006400'
BLUE = '1E90FF'
BLACK = '000000'
LIGHT_GREEN = '90EE90'
LIGHT_BLUE = 'B7DEE8'
LIGHT_GRAY = 'D3D3D3'
PINK = 'FFD9D9'
WHITE_SMOKE = 'F5F5F5'
WHITE = 'FFFFFF'
BLACK = '000000'
LILA = 'C8A2C8'


def _styleCell(cell, color, fontsize, textColor, bold):
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    font = Font(bold=bold, size=fontsize, color=textColor)
    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    align = Alignment(horizontal='center', vertical='center')
    cell.fill = fill
    cell.font = font
    cell.border = border
    cell.alignment = align


def _styleNormal(cell):
    _styleCell(cell, WHITE_SMOKE, 12, BLACK, False)
    return cell


def _styleHeaderNormal(cell):
    _styleCell(cell, LILA, 14, BLACK, True)
    return cell


def _styleHeaderGreen(cell):
    _styleCell(cell, LIGHT_GREEN, 14, BLACK, True)
    return cell


def _styleHeaderRed(cell):
    _styleCell(cell, LIGHT_BLUE, 14, BLACK, True)
    return cell
