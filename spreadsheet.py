import openpyxl


class Spreadsheet:
    def __init__(self, name: str):
        self.name = name
        self.book = openpyxl.Workbook()

        try:
            self.book = openpyxl.load_workbook(self.name)
        except FileNotFoundError:
            self.users_sheet = self.book.create_sheet("Usuarios")
            self.works_sheet = self.book.create_sheet("Obras")
            self.jobs_sheet = self.book.create_sheet("Trabajos")
            self.users_sheet.append(["IdUsuario"])
            self.works_sheet.append(["Nombre"])
            self.jobs_sheet.append(
                ["NombreObra", "IdUsuario", "Fecha", "Función", "Capitulo", "Servidor", "Canal"])
            self.book.save(self.name)

        self.users_sheet = self.book["Usuarios"]
        self.works_sheet = self.book["Obras"]
        self.jobs_sheet = self.book["Trabajos"]

    def get_users(self):
        users = []
        for row in self.users_sheet.iter_rows(values_only=True):
            users.append(row[0])
        return users

    def get_works(self):
        works = []
        for row in self.works_sheet.iter_rows(values_only=True):
            works.append(row[0])
        return works

    def get_jobs(self):
        jobs = []
        for row in self.jobs_sheet.iter_rows(values_only=True):
            jobs.append(row)
        return jobs

    def add_user(self, user_id):
        self.users_sheet.append([user_id])
        self.save()

    def add_work(self, work_name):
        self.works_sheet.append([work_name])
        self.save()

    def add_job(self, work_name, user_id, date, function, chapter, server, channel):
        self.jobs_sheet.append(
            [work_name, user_id, date, function, chapter, server, channel])
        self.save()

    def remove_user(self, user_id):
        for row in self.users_sheet.iter_rows():
            if row[0].value == user_id:
                self.users_sheet.delete_rows(row[0].row)
                break

    def remove_work(self, work_name):
        for row in self.works_sheet.iter_rows():
            if row[0].value == work_name:
                self.works_sheet.delete_rows(row[0].row)
                break

    def save(self):
        self.book.save()

    def close(self):
        self.save()
        self.book.close()
